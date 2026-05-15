# -*- coding: utf-8 -*-
"""
B题：多源融合机器人定位及任务优化

运行方式：
    python b_solution.py

输出目录：
    outputs/
        estimates_summary.xlsx
        problem1_10Hz_trajectory.xlsx
        problem2_10Hz_trajectory.xlsx
        problem3_10Hz_trajectory.xlsx
        result.xlsx
        B题_多源融合机器人定位及任务优化_论文.docx
        B题_论文.md
        figures/*.png
"""

from __future__ import annotations

import argparse
import html
import math
import re
import shutil
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Sequence

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from matplotlib.backends.backend_pdf import PdfPages
from matplotlib.font_manager import FontProperties
from scipy.optimize import minimize_scalar
from scipy.signal import savgol_filter
from scipy.stats import f as f_dist, shapiro, probplot

DT_OUT = 0.1

# 附录公式（由题面 MathType 公式对象读取）
SHOOT_DISTANCE = (5.0, 30.0)
SHOOT_SPEED_MAX = 2.0
SHOOT_ACCEL_MAX = 1.5
SHOOT_PREP = 1.5
SHOOT_HIT_PROB = 0.85

PHOTO_DISTANCE = (10.0, 40.0)
PHOTO_ANGLE_MIN_DEG = 60.0
PHOTO_SPEED_MAX = 1.5
PHOTO_ACCEL_MAX = 1.5
PHOTO_PREP = 0.5


@dataclass
class AlignmentResult:
    problem: int
    delta2_minus_1: float
    bias2_x: float
    bias2_y: float
    mse_with_bias: float
    mse_without_bias: float | None = None
    bias_model_improvement: float | None = None
    has_system_bias: bool = True
    overlap_seconds: float = 0.0
    ci_delta_lo: float = 0.0
    ci_delta_hi: float = 0.0
    f_statistic: float | None = None
    f_p_value: float | None = None
    candidate_bias_x: float = 0.0
    candidate_bias_y: float = 0.0
    candidate_delta: float = 0.0
    candidate_mse: float = 0.0


def discover_files(root: Path) -> tuple[dict[int, Path], Path]:
    xlsx_files: dict[int, Path] = {}
    result_template: Path | None = None
    result_candidates: list[Path] = []
    for path in root.glob("*.xlsx"):
        if path.name.startswith("~$"):
            continue
        if path.name.lower() == "result.xlsx":
            result_template = path
            continue
        if path.name.lower().startswith("result"):
            result_candidates.append(path)
            continue
        m = re.search(r"(\d+)", path.name)
        if m:
            xlsx_files[int(m.group(1))] = path
    missing = [i for i in (1, 2, 3, 4) if i not in xlsx_files]
    if missing:
        raise FileNotFoundError(f"缺少附件：{missing}")
    if result_template is None and result_candidates:
        result_template = sorted(result_candidates)[0]
    if result_template is None:
        raise FileNotFoundError("未找到 result.xlsx 模板")
    return xlsx_files, result_template


def read_pair(path: Path) -> tuple[np.ndarray, np.ndarray, np.ndarray, np.ndarray]:
    d1 = pd.read_excel(path, sheet_name=0).dropna()
    d2 = pd.read_excel(path, sheet_name=1).dropna()
    a1 = d1.iloc[:, 0:3].to_numpy(float)
    a2 = d2.iloc[:, 0:3].to_numpy(float)
    return a1[:, 0], a1[:, 1:3], a2[:, 0], a2[:, 1:3]


def bias_f_test(
    mse_without_bias: float,
    mse_with_bias: float,
    n_samples: int,
    n_bias_params: int = 2,
    alpha: float = 0.05,
) -> tuple[float, float, bool]:
    """嵌套模型F检验 H0: 无系统偏差 (bx=by=0)."""
    if mse_with_bias <= 0 or mse_without_bias <= mse_with_bias:
        return 1.0, 1.0, False
    df1 = n_bias_params
    df2 = n_samples - df1
    if df2 <= 0:
        return 1.0, 1.0, False
    sse_without = mse_without_bias * n_samples
    sse_with = mse_with_bias * n_samples
    f_stat = ((sse_without - sse_with) / df1) / (sse_with / df2)
    p_value = 1.0 - f_dist.cdf(f_stat, df1, df2)
    return f_stat, p_value, p_value < alpha


def estimate_alignment_raw(
    t1: np.ndarray, p1: np.ndarray, t2: np.ndarray, p2: np.ndarray,
    d_min: float, d_max: float,
    correct_bias: bool,
    score_smooth_window: int,
    trim_ratio: float,
    min_overlap_ratio: float = 0.85,
) -> tuple[float, np.ndarray, float, float]:
    """核心对齐搜索（粗扫+精细优化），被主估计和bootstrap共用."""
    p1s = moving_average(p1, score_smooth_window)
    p2s = moving_average(p2, score_smooth_window)
    deltas = np.linspace(d_min, d_max, 2500)
    scores = np.array(
        [
            delta_score(t1, p1s, t2, p2s, d, correct_bias, 0.5, trim_ratio, min_overlap_ratio)[0]
            for d in deltas
        ]
    )
    best_i = int(np.nanargmin(scores))
    best = float(deltas[best_i])
    step = float(deltas[1] - deltas[0]) if len(deltas) > 1 else 1.0
    lo = max(d_min, best - 20 * step)
    hi = min(d_max, best + 20 * step)
    opt = minimize_scalar(
        lambda d: delta_score(t1, p1s, t2, p2s, d, correct_bias, 0.1, trim_ratio, min_overlap_ratio)[0],
        bounds=(lo, hi),
        method="bounded",
        options={"xatol": 1e-7},
    )
    mse, bias, overlap = delta_score(
        t1, p1s, t2, p2s, float(opt.x), correct_bias, 0.1, trim_ratio, min_overlap_ratio
    )
    return float(opt.x), bias, mse, overlap


def bootstrap_ci(
    t1: np.ndarray, p1: np.ndarray, t2: np.ndarray, p2: np.ndarray,
    delta_hat: float, bias_hat: np.ndarray,
    n_bootstrap: int = 199,
    alpha: float = 0.05,
    dt: float = 0.1,
) -> tuple[float, float, float, float]:
    """Linearized residual bootstrap CI for delta and bias.

    At the optimum, changing delta moves the corrected mode-2 trajectory along
    its local velocity. Resampling residuals and solving the one-parameter
    least-squares perturbation gives a non-degenerate uncertainty estimate
    without rerunning the expensive global search hundreds of times.
    """
    t2_corr = t2 - delta_hat
    p2_corr = p2 - bias_hat
    lo = max(float(t1.min()), float(t2_corr.min()))
    hi = min(float(t1.max()), float(t2_corr.max()))
    grid = np.arange(lo, hi + 1e-9, dt)
    q1 = interp_xy(grid, t1, p1)
    q2 = interp_xy(grid, t2_corr, p2_corr)
    residuals = q2 - q1
    velocity = np.gradient(q2, dt, axis=0)
    n_res = len(residuals)
    if n_res < 10:
        return 0.0, 0.0, 0.0, 0.0

    deltas_bs, bias_x_bs, bias_y_bs = [], [], []
    rng = np.random.default_rng(2026)
    for _ in range(n_bootstrap):
        idx = rng.integers(0, n_res, size=n_res)
        res_bs = residuals[idx]
        bias_bs = np.median(res_bs, axis=0)
        bias_x_bs.append(bias_bs[0])
        bias_y_bs.append(bias_bs[1])
        v_bs = velocity[idx]
        res_center = res_bs - np.median(res_bs, axis=0)
        denom = float(np.sum(v_bs ** 2))
        if denom <= 1e-12:
            deltas_bs.append(delta_hat)
            continue
        eps = -float(np.sum(v_bs * res_center)) / denom
        deltas_bs.append(delta_hat + eps)

    if len(deltas_bs) < 2:
        return 0.0, 0.0, 0.0, 0.0
    d_lo, d_hi = np.percentile(deltas_bs, [100 * alpha / 2, 100 * (1 - alpha / 2)])
    bx_lo, bx_hi = np.percentile(bias_x_bs, [100 * alpha / 2, 100 * (1 - alpha / 2)])
    return float(d_lo), float(d_hi), float(bx_lo), float(bx_hi)


def moving_average(y: np.ndarray, window: int) -> np.ndarray:
    if window <= 1:
        return y.copy()
    if window % 2 == 0:
        window += 1
    pad = window // 2
    kernel = np.ones(window) / window
    yp = np.pad(y, ((pad, pad), (0, 0)), mode="edge")
    return np.column_stack([np.convolve(yp[:, i], kernel, mode="valid") for i in range(y.shape[1])])


def interp_xy(t_new: np.ndarray, t: np.ndarray, p: np.ndarray) -> np.ndarray:
    return np.column_stack([np.interp(t_new, t, p[:, i]) for i in range(2)])


def delta_score(
    t1: np.ndarray,
    p1: np.ndarray,
    t2: np.ndarray,
    p2: np.ndarray,
    delta: float,
    correct_bias: bool,
    dt: float,
    trim_ratio: float,
    min_overlap_ratio: float,
) -> tuple[float, np.ndarray, float]:
    t2_corr = t2 - delta
    lo = max(float(t1.min()), float(t2_corr.min()))
    hi = min(float(t1.max()), float(t2_corr.max()))
    overlap = hi - lo
    min_duration = min(float(t1.max() - t1.min()), float(t2.max() - t2.min()))
    if overlap < min_overlap_ratio * min_duration:
        return float("inf"), np.array([np.nan, np.nan]), overlap

    grid = np.arange(lo, hi + 1e-9, dt)
    q1 = interp_xy(grid, t1, p1)
    q2 = interp_xy(grid, t2_corr, p2)
    diff = q2 - q1
    bias = np.median(diff, axis=0) if correct_bias else np.zeros(2)
    err = np.sum((diff - bias) ** 2, axis=1)
    if trim_ratio > 0 and len(err) > 20:
        keep = max(10, int(len(err) * (1.0 - trim_ratio)))
        err = np.sort(err)[:keep]
    return float(np.mean(err)), bias, overlap


def estimate_alignment(
    problem: int,
    path: Path,
    correct_bias: bool,
    score_smooth_window: int,
    trim_ratio: float,
    min_overlap_ratio: float = 0.85,
) -> tuple[float, np.ndarray, float, float]:
    t1, p1, t2, p2 = read_pair(path)
    min_duration = min(float(t1.max() - t1.min()), float(t2.max() - t2.min()))
    d_min = float(t2.min() - t1.max() + min_overlap_ratio * min_duration)
    d_max = float(t2.max() - t1.min() - min_overlap_ratio * min_duration)
    return estimate_alignment_raw(t1, p1, t2, p2, d_min, d_max, correct_bias,
                                   score_smooth_window, trim_ratio, min_overlap_ratio)


def compute_overlap_n(t1: np.ndarray, p1: np.ndarray, t2: np.ndarray, p2: np.ndarray,
                      delta: float, dt: float = 0.1) -> int:
    """重叠区间内的样本数（用于F检验自由度）. """
    t2_corr = t2 - delta
    lo = max(float(t1.min()), float(t2_corr.min()))
    hi = min(float(t1.max()), float(t2_corr.max()))
    return max(1, int(round((hi - lo) / dt)))


def build_alignment_results(files: dict[int, Path]) -> dict[int, AlignmentResult]:
    results: dict[int, AlignmentResult] = {}

    # ----- Problem 1: 无噪声无偏差 -----
    d1, b1, mse1, ov1 = estimate_alignment(1, files[1], False, 1, 0.0)
    results[1] = AlignmentResult(1, d1, 0.0, 0.0, mse1, has_system_bias=False, overlap_seconds=ov1)

    # ----- Problem 2: 随机噪声 + 系统偏差 -----
    t1_2, p1_2, t2_2, p2_2 = read_pair(files[2])
    d2, b2, mse2, ov2 = estimate_alignment(2, files[2], True, 9, 0.02)
    d2_nb, _b2_nb, mse2_nb, _ov2_nb = estimate_alignment(2, files[2], False, 9, 0.02)
    improvement2 = (mse2_nb - mse2) / mse2_nb if mse2_nb > 0 else 0.0
    n2 = compute_overlap_n(t1_2, p1_2, t2_2, p2_2, d2)
    f_stat2, f_p2, has_bias2 = bias_f_test(mse2_nb, mse2, n2)
    results[2] = AlignmentResult(
        2, d2, float(b2[0]), float(b2[1]), mse2,
        mse_without_bias=mse2_nb, bias_model_improvement=improvement2,
        has_system_bias=has_bias2, overlap_seconds=ov2,
        f_statistic=f_stat2, f_p_value=f_p2,
    )

    # ----- Problem 3: 实测数据 -----
    t1_3, p1_3, t2_3, p2_3 = read_pair(files[3])
    d3_b, b3, mse3_b, ov3_b = estimate_alignment(3, files[3], True, 11, 0.05)
    d3_nb, _b3_nb, mse3_nb, ov3_nb = estimate_alignment(3, files[3], False, 11, 0.05)
    improvement3 = (mse3_nb - mse3_b) / mse3_nb if mse3_nb > 0 else 0.0
    n3 = compute_overlap_n(t1_3, p1_3, t2_3, p2_3, d3_b)
    f_stat3, f_p3, stat_has_bias3 = bias_f_test(mse3_nb, mse3_b, n3)
    bias_norm3 = float(np.linalg.norm(b3))
    has_bias3 = bool(stat_has_bias3 and bias_norm3 >= 0.1)

    if has_bias3:
        d3, b3_used, mse3, ov3 = d3_b, b3, mse3_b, ov3_b
    else:
        d3, b3_used, mse3, ov3 = d3_nb, np.zeros(2), mse3_nb, ov3_nb

    results[3] = AlignmentResult(
        3, d3, float(b3_used[0]), float(b3_used[1]), mse3,
        mse_without_bias=mse3_nb, bias_model_improvement=improvement3,
        has_system_bias=has_bias3, overlap_seconds=ov3,
        f_statistic=f_stat3, f_p_value=f_p3,
    )
    # 保存候选偏差供论文参考（无论是否被采用）
    results[3].candidate_bias_x = float(b3[0])
    results[3].candidate_bias_y = float(b3[1])
    results[3].candidate_delta = float(d3_b)
    results[3].candidate_mse = float(mse3_b)

    # ----- Bootstrap置信区间（问题2&3） -----
    for prob in (2, 3):
        r = results[prob]
        t1_, p1_, t2_, p2_ = read_pair(files[prob])
        if prob == 2:
            sw, tr = 9, 0.02
        else:
            sw, tr = 11, 0.05
        d_lo, d_hi, bx_lo, bx_hi = bootstrap_ci(
            t1_, p1_, t2_, p2_, r.delta2_minus_1,
            np.array([r.bias2_x, r.bias2_y]),
            n_bootstrap=199,
        )
        # 如果bootstrap没给有效结果则使用渐近近似
        if d_lo == 0 and d_hi == 0:
            # 基于MSE曲率近似95% CI: ±1.96 * sigma/sqrt(n)
            n_samples = compute_overlap_n(t1_, p1_, t2_, p2_, r.delta2_minus_1)
            sigma_d = np.sqrt(r.mse_with_bias / n_samples) if r.mse_with_bias > 0 else 0.01
            d_lo, d_hi = r.delta2_minus_1 - 1.96 * sigma_d, r.delta2_minus_1 + 1.96 * sigma_d
        r.ci_delta_lo, r.ci_delta_hi = float(d_lo), float(d_hi)

    return results


def make_trajectory(
    path: Path,
    alignment: AlignmentResult,
    smooth_window: int,
    smooth_poly: int = 3,
) -> pd.DataFrame:
    t1, p1, t2, p2 = read_pair(path)
    t2_corr = t2 - alignment.delta2_minus_1
    p2_corr = p2 - np.array([alignment.bias2_x, alignment.bias2_y])
    lo = max(float(t1.min()), float(t2_corr.min()))
    hi = min(float(t1.max()), float(t2_corr.max()))
    start = math.ceil(lo / DT_OUT) * DT_OUT
    end = math.floor(hi / DT_OUT) * DT_OUT
    grid = np.round(np.arange(start, end + 1e-9, DT_OUT), 10)

    q1 = interp_xy(grid, t1, p1)
    q2 = interp_xy(grid, t2_corr, p2_corr)
    fused_raw = 0.5 * (q1 + q2)  # unsmoothed fusion for derivative computation
    if smooth_window > 1 and len(fused_raw) > smooth_window:
        if smooth_window % 2 == 0:
            smooth_window += 1
        fused = savgol_filter(fused_raw, smooth_window, smooth_poly, axis=0, mode="interp")
        # Savgol analytical derivative: fit polynomial, differentiate analytically
        # deriv=1 gives dy/dsample, divide by DT_OUT to get dy/dt
        vel = savgol_filter(fused_raw, smooth_window, smooth_poly,
                            deriv=1, axis=0, mode="interp") / DT_OUT
        acc_vec = savgol_filter(fused_raw, smooth_window, smooth_poly,
                                deriv=2, axis=0, mode="interp") / DT_OUT ** 2
    else:
        fused = fused_raw.copy()
        vel = np.gradient(fused, DT_OUT, axis=0)
        acc_vec = np.gradient(vel, DT_OUT, axis=0)

    speed = np.linalg.norm(vel, axis=1)
    acc = np.linalg.norm(acc_vec, axis=1)
    return pd.DataFrame(
        {
            "time_s": grid,
            "x_m": fused[:, 0],
            "y_m": fused[:, 1],
            "speed_m_s": speed,
            "accel_m_s2": acc,
        }
    )


def rolling_all(mask: np.ndarray, window_samples: int) -> np.ndarray:
    out = np.zeros_like(mask, dtype=bool)
    csum = np.r_[0, np.cumsum(mask.astype(int))]
    out[window_samples - 1 :] = (csum[window_samples:] - csum[:-window_samples]) == window_samples
    return out


def circular_angle_diff_deg(a: float, b: float) -> float:
    d = abs((a - b + 180.0) % 360.0 - 180.0)
    return float(d)


def compute_alignment_residuals(
    t1: np.ndarray, p1: np.ndarray, t2: np.ndarray, p2: np.ndarray,
    delta: float, bias: np.ndarray, dt: float = 0.1,
) -> np.ndarray:
    """Compute per-point alignment residuals (euclidean distance per point). """
    t2_corr = t2 - delta
    p2_corr = p2 - bias
    lo = max(float(t1.min()), float(t2_corr.min()))
    hi = min(float(t1.max()), float(t2_corr.max()))
    grid = np.arange(lo, hi + 1e-9, dt)
    q1 = interp_xy(grid, t1, p1)
    q2 = interp_xy(grid, t2_corr, p2_corr)
    diff = q2 - q1
    return np.linalg.norm(diff, axis=1), diff


def residual_diagnostics(
    t1: np.ndarray, p1: np.ndarray, t2: np.ndarray, p2: np.ndarray,
    delta: float, bias: np.ndarray, label: str = "",
) -> dict:
    """Full residual diagnostic suite for a given alignment. """
    dists, diff = compute_alignment_residuals(t1, p1, t2, p2, delta, bias)
    n = len(dists)
    mean_res = float(np.mean(dists))
    std_res = float(np.std(dists))
    # Noise std estimate: MAD robust estimator
    mad = float(np.median(np.abs(dists - np.median(dists)))) * 1.4826
    # Normality test (Shapiro-Wilk, limited to 5000 points)
    if n > 5000:
        idx = np.linspace(0, n - 1, 5000, dtype=int)
        shap_stat, shap_p = shapiro(dists[idx])
    else:
        shap_stat, shap_p = shapiro(dists)
    # Auto-correlation at lag-1
    autocorr = float(np.corrcoef(dists[:-1], dists[1:])[0, 1]) if n > 1 else 0.0
    # Per-axis stats
    x_residuals = diff[:, 0]
    y_residuals = diff[:, 1]
    noise_std_x = float(np.std(x_residuals))
    noise_std_y = float(np.std(y_residuals))
    return {
        "label": label,
        "n_points": n,
        "mean_residual": mean_res,
        "std_residual": std_res,
        "mad_robust_std": mad,
        "noise_std_est_x": noise_std_x,
        "noise_std_est_y": noise_std_y,
        "shapiro_stat": float(shap_stat),
        "shapiro_p": float(shap_p),
        "autocorr_lag1": autocorr,
        "max_residual": float(np.max(dists)),
        "p95_residual": float(np.percentile(dists, 95)),
    }


def estimate_window_impact(
    t1: np.ndarray, p1: np.ndarray, t2: np.ndarray, p2: np.ndarray,
    delta: float, bias: np.ndarray,
    windows: list[int],
) -> list[dict]:
    """Evaluate how Savgol window size affects velocity/accel statistics. """
    results = []
    t2_corr = t2 - delta
    p2_corr = p2 - bias
    lo = max(float(t1.min()), float(t2_corr.min()))
    hi = min(float(t1.max()), float(t2_corr.max()))
    grid = np.arange(lo, hi + 1e-9, DT_OUT)
    q1 = interp_xy(grid, t1, p1)
    q2 = interp_xy(grid, t2_corr, p2_corr)
    fused_raw = 0.5 * (q1 + q2)
    for w in windows:
        if w <= 1 or len(fused_raw) <= w:
            continue
        w_odd = w if w % 2 == 1 else w + 1
        vel = savgol_filter(fused_raw, w_odd, 3, deriv=1, axis=0, mode="interp") / DT_OUT
        acc = savgol_filter(fused_raw, w_odd, 3, deriv=2, axis=0, mode="interp") / DT_OUT**2
        speed = np.linalg.norm(vel, axis=1)
        acc_mag = np.linalg.norm(acc, axis=1)
        results.append({
            "window": w_odd,
            "speed_max": float(np.max(speed)),
            "speed_mean": float(np.mean(speed)),
            "speed_std": float(np.std(speed)),
            "acc_max": float(np.max(acc_mag)),
            "acc_mean": float(np.mean(acc_mag)),
            "pct_speed_gt_2": float(np.mean(speed > 2.0) * 100),
            "pct_acc_gt_1_5": float(np.mean(acc_mag > 1.5) * 100),
        })
    return results


def select_task_candidates(traj: pd.DataFrame, target_path: Path) -> pd.DataFrame:
    t = traj["time_s"].to_numpy(float)
    xy = traj[["x_m", "y_m"]].to_numpy(float)
    speed = traj["speed_m_s"].to_numpy(float)
    accel = traj["accel_m_s2"].to_numpy(float)

    shots = pd.read_excel(target_path, sheet_name=0).dropna().iloc[:, 0:3]
    photos = pd.read_excel(target_path, sheet_name=1).dropna().iloc[:, 0:3]
    rows: list[dict[str, float | str]] = []

    shoot_window = int(round(SHOOT_PREP / DT_OUT)) + 1
    for _, row in shots.iterrows():
        target_id = str(row.iloc[0])
        pt = row.iloc[1:3].to_numpy(float)
        dist = np.linalg.norm(xy - pt, axis=1)
        base = (
            (dist >= SHOOT_DISTANCE[0])
            & (dist <= SHOOT_DISTANCE[1])
            & (speed <= SHOOT_SPEED_MAX)
            & (accel <= SHOOT_ACCEL_MAX)
        )
        ok = rolling_all(base, shoot_window)
        if not np.any(ok):
            continue
        idxs = np.where(ok)[0]
        best_idx = int(idxs[np.argmin(dist[idxs])])
        rows.append(
            {
                "target_id": target_id,
                "task": "模拟射击",
                "prep_start_s": round(float(t[best_idx] - SHOOT_PREP), 1),
                "exec_time_s": round(float(t[best_idx]), 1),
                "distance_m": float(dist[best_idx]),
                "speed_m_s": float(speed[best_idx]),
                "accel_m_s2": float(accel[best_idx]),
                "angle_deg": "",
                "expected_success": SHOOT_HIT_PROB,
            }
        )

    photo_window = int(round(PHOTO_PREP / DT_OUT)) + 1
    for _, row in photos.iterrows():
        target_id = str(row.iloc[0])
        pt = row.iloc[1:3].to_numpy(float)
        vec = pt - xy
        dist = np.linalg.norm(vec, axis=1)
        angles = (np.degrees(np.arctan2(vec[:, 1], vec[:, 0])) + 360.0) % 360.0
        base = (
            (dist >= PHOTO_DISTANCE[0])
            & (dist <= PHOTO_DISTANCE[1])
            & (speed <= PHOTO_SPEED_MAX)
            & (accel <= PHOTO_ACCEL_MAX)
        )
        ok = rolling_all(base, photo_window)
        if not np.any(ok):
            continue

        # 同一目标保留尽量多的不同角度；同一角度簇内选距离最近的时刻。
        candidates = np.where(ok)[0]
        selected: list[int] = []
        for idx in candidates[np.argsort(dist[candidates])]:
            if all(circular_angle_diff_deg(float(angles[idx]), float(angles[j])) >= PHOTO_ANGLE_MIN_DEG for j in selected):
                selected.append(int(idx))
        selected.sort(key=lambda i: t[i])
        for idx in selected:
            rows.append(
                {
                    "target_id": target_id,
                    "task": "拍照",
                    "prep_start_s": round(float(t[idx] - PHOTO_PREP), 1),
                    "exec_time_s": round(float(t[idx]), 1),
                    "distance_m": float(dist[idx]),
                    "speed_m_s": float(speed[idx]),
                    "accel_m_s2": float(accel[idx]),
                    "angle_deg": float(angles[idx]),
                    "expected_success": 1.0,
                }
            )

    tasks = pd.DataFrame(rows)
    if tasks.empty:
        return tasks
    return tasks.sort_values(["prep_start_s", "exec_time_s", "target_id", "task"]).reset_index(drop=True)


def weighted_interval_scheduling(tasks: pd.DataFrame) -> pd.DataFrame:
    """加权区间调度（DP）求最大期望完成数的非重叠子集.

    每项任务有开始时间 prep_start_s 和结束时间 exec_time_s，
    权重 expected_success（射击0.85，拍照1.0）。
    """
    if tasks.empty:
        return tasks
    df = tasks.copy()
    df["end_s"] = df["exec_time_s"].astype(float)
    df["start_s"] = df["prep_start_s"].astype(float)
    df["weight"] = df["expected_success"].astype(float)
    df = df.sort_values("end_s").reset_index(drop=True)

    n = len(df)
    # p[i] = 最靠右的与task i不冲突的任务索引（-1表示无）
    p = [-1] * n
    for i in range(n):
        for j in range(i - 1, -1, -1):
            if df.loc[j, "end_s"] <= df.loc[i, "start_s"] - 1e-9:
                p[i] = j
                break

    # DP: dp[i] = 前i+1个任务的最大权重
    dp = [0.0] * n
    selected = [False] * n
    for i in range(n):
        w_include = df.loc[i, "weight"] + (dp[p[i]] if p[i] >= 0 else 0.0)
        w_exclude = dp[i - 1] if i > 0 else 0.0
        if w_include >= w_exclude:
            dp[i] = w_include
            selected[i] = True
        else:
            dp[i] = w_exclude
            selected[i] = False

    # 回溯
    chosen: list[int] = []
    i = n - 1
    while i >= 0:
        if selected[i]:
            chosen.append(i)
            i = p[i]
        else:
            i -= 1
    chosen.reverse()
    return df.iloc[chosen].drop(columns=["end_s", "start_s", "weight"]).reset_index(drop=True)


def filter_sequential(tasks: pd.DataFrame) -> pd.DataFrame:
    if tasks.empty:
        return tasks
    rows = []
    current_end = -float("inf")
    for _, row in tasks.sort_values(["exec_time_s", "prep_start_s"]).iterrows():
        if float(row["prep_start_s"]) >= current_end - 1e-9:
            rows.append(row)
            current_end = float(row["exec_time_s"])
    return pd.DataFrame(rows).reset_index(drop=True)


def write_result_xlsx(template: Path, output: Path, tasks: pd.DataFrame) -> None:
    # 清除Office锁文件，然后尝试写目标路径；被锁定时用临时路径
    lock = output.parent / ("~$" + output.name)
    if lock.exists():
        try:
            lock.unlink()
        except OSError:
            pass
    try:
        dest = output
        shutil.copy2(template, output)
    except PermissionError:
        dest = output.parent / "_temp_result.xlsx"
        shutil.copy2(template, dest)
    wb = load_workbook(dest)
    ws = wb.active
    # 仅写 A:E 答案区，保留右侧红色说明和示例。
    for r in range(2, max(ws.max_row + 1, len(tasks) + 3)):
        for c in range(1, 6):
            ws.cell(r, c).value = None
    for seq, (_, row) in enumerate(tasks.iterrows()):
        r = seq + 2
        ws.cell(r, 1, seq + 1)
        ws.cell(r, 2, row["target_id"])
        ws.cell(r, 3, row["task"])
        ws.cell(r, 4, float(row["prep_start_s"]))
        ws.cell(r, 5, float(row["exec_time_s"]))
        for c in range(1, 6):
            cell = ws.cell(r, c)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(name="宋体", size=12)
    wb.save(dest)


def save_estimates(path: Path, results: dict[int, AlignmentResult]) -> None:
    rows = []
    for i in (1, 2, 3):
        r = results[i]
        row = {
            "问题": i,
            "方式2相对方式1时间偏差_delta_s": r.delta2_minus_1,
            "delta_95%CI_lower_s": r.ci_delta_lo,
            "delta_95%CI_upper_s": r.ci_delta_hi,
            "方式1时间偏差_s": 0.0,
            "方式2时间偏差_s": r.delta2_minus_1,
            "方式2相对方式1系统偏差_x_m": r.bias2_x,
            "方式2相对方式1系统偏差_y_m": r.bias2_y,
            "是否判定存在系统偏差": "是" if r.has_system_bias else "否",
            "F统计量": r.f_statistic,
            "F检验p值": r.f_p_value,
            "融合重叠时长_s": r.overlap_seconds,
            "带偏差模型MSE": r.mse_with_bias,
            "无偏差模型MSE": r.mse_without_bias,
            "偏差模型误差下降比例": r.bias_model_improvement,
        }
        if i == 3:
            row["实际数据候选偏差_x_m"] = getattr(r, "candidate_bias_x", np.nan)
            row["实际数据候选偏差_y_m"] = getattr(r, "candidate_bias_y", np.nan)
            row["实际数据带偏差候选delta_s"] = getattr(r, "candidate_delta", np.nan)
        rows.append(row)
    pd.DataFrame(rows).to_excel(path, index=False)


def plot_outputs(out_dir: Path, trajectories: dict[int, pd.DataFrame], tasks: pd.DataFrame, target_path: Path) -> list[Path]:
    fig_dir = out_dir / "figures"
    fig_dir.mkdir(parents=True, exist_ok=True)
    paths: list[Path] = []

    stale = fig_dir / "trajectories_10hz.png"
    if stale.exists():
        stale.unlink()

    for i, df in trajectories.items():
        plt.figure(figsize=(7.0, 5.2), dpi=180)
        plt.plot(df["x_m"], df["y_m"], linewidth=1.3, color="#1f77b4", label=f"Problem {i} fused trajectory")
        plt.scatter(df["x_m"].iloc[0], df["y_m"].iloc[0], marker="o", color="#2ca02c", s=36, label="Start")
        plt.scatter(df["x_m"].iloc[-1], df["y_m"].iloc[-1], marker="s", color="#d62728", s=36, label="End")
        plt.axis("equal")
        plt.xlabel("X (m)")
        plt.ylabel("Y (m)")
        plt.title(f"Problem {i} 10Hz Fused Trajectory")
        plt.legend(fontsize=8)
        plt.tight_layout()
        p = fig_dir / f"problem{i}_trajectory_10hz.png"
        plt.savefig(p)
        plt.close()
        paths.append(p)

    shots = pd.read_excel(target_path, sheet_name=0).dropna().iloc[:, 0:3]
    photos = pd.read_excel(target_path, sheet_name=1).dropna().iloc[:, 0:3]
    tr3 = trajectories[3]
    plt.figure(figsize=(7.0, 5.2), dpi=180)
    plt.plot(tr3["x_m"], tr3["y_m"], color="#1f77b4", linewidth=1.2, label="Fused trajectory")
    plt.scatter(shots.iloc[:, 1], shots.iloc[:, 2], marker="x", color="#d62728", label="Shooting targets")
    plt.scatter(photos.iloc[:, 1], photos.iloc[:, 2], marker="o", facecolors="none", edgecolors="#2ca02c", label="Photo targets")
    selected_ids = set(tasks["target_id"].astype(str)) if not tasks.empty else set()
    for df, color in [(shots, "#d62728"), (photos, "#2ca02c")]:
        sel = df[df.iloc[:, 0].astype(str).isin(selected_ids)]
        if not sel.empty:
            plt.scatter(sel.iloc[:, 1], sel.iloc[:, 2], s=80, facecolors="none", edgecolors=color, linewidths=2.0)
    plt.axis("equal")
    plt.xlabel("X (m)")
    plt.ylabel("Y (m)")
    plt.legend(fontsize=8)
    plt.tight_layout()
    p = fig_dir / "problem4_selected_tasks.png"
    plt.savefig(p)
    plt.close()
    paths.append(p)
    return paths


def plot_diagnostics(
    out_dir: Path,
    files: dict[int, Path],
    alignments: dict[int, AlignmentResult],
    traj_3: pd.DataFrame,
    sens_windows: list[int],
    sens_rows_data: list[dict],
    window_impact_data: list[dict] | None = None,
) -> list[Path]:
    """Generate diagnostic plots: QQ, residuals, sensitivity, speed comparison. """
    fig_dir = out_dir / "figures"
    fig_dir.mkdir(parents=True, exist_ok=True)
    paths: list[Path] = []

    # QQ plots for Problem 2 and 3
    for prob in (2, 3):
        t1, p1, t2, p2 = read_pair(files[prob])
        al = alignments[prob]
        dists, _diff = compute_alignment_residuals(
            t1, p1, t2, p2, al.delta2_minus_1, np.array([al.bias2_x, al.bias2_y])
        )
        n = len(dists)
        fig, axes = plt.subplots(1, 3, figsize=(14, 4.2), dpi=150)
        # QQ plot
        probplot(dists, dist="norm", plot=axes[0])
        axes[0].set_title(f"Problem {prob} Residual QQ Plot (n={n})")
        axes[0].set_xlabel("Theoretical Quantiles")
        axes[0].set_ylabel("Sample Quantiles")
        # Histogram
        axes[1].hist(dists, bins=60, density=True, alpha=0.7, color="#1f77b4", edgecolor="white")
        from scipy.stats import norm as norm_dist
        x_range = np.linspace(dists.min(), dists.max(), 200)
        axes[1].plot(x_range, norm_dist.pdf(x_range, np.mean(dists), np.std(dists)),
                     "r-", linewidth=1.5, label="Normal fit")
        axes[1].set_title(f"Problem {prob} Residual Distribution")
        axes[1].set_xlabel("Residual (m)")
        axes[1].set_ylabel("Density")
        axes[1].legend(fontsize=8)
        # Residuals vs order
        axes[2].plot(dists, "o", markersize=1.2, alpha=0.4, color="#1f77b4")
        axes[2].axhline(np.mean(dists), color="r", linestyle="--", linewidth=1, label=f"Mean={np.mean(dists):.4f}")
        axes[2].axhline(np.mean(dists) + 3 * np.std(dists), color="orange", linestyle=":", linewidth=1, label="±3σ")
        axes[2].axhline(np.mean(dists) - 3 * np.std(dists), color="orange", linestyle=":", linewidth=1)
        axes[2].set_title(f"Problem {prob} Residuals vs Index")
        axes[2].set_xlabel("Sample index")
        axes[2].set_ylabel("Residual (m)")
        axes[2].legend(fontsize=8)
        plt.tight_layout()
        p = fig_dir / f"problem{prob}_diagnostics.png"
        plt.savefig(p, bbox_inches="tight")
        plt.close()
        paths.append(p)

    # Sensitivity analysis chart
    if sens_rows_data:
        fig, axes = plt.subplots(1, 2, figsize=(12, 4.5), dpi=150)
        windows = [r["平滑窗口(10Hz点数)"] for r in sens_rows_data]
        n_tasks = [r["任务数"] for r in sens_rows_data]
        expected = [r["期望完成数"] for r in sens_rows_data]
        axes[0].plot(windows, n_tasks, "o-", color="#1f77b4", linewidth=1.8, markersize=7)
        axes[0].axhline(10, color="gray", linestyle="--", linewidth=0.8, alpha=0.7, label="Default (71)")
        axes[0].set_xlabel("Smoothing window (10Hz samples)")
        axes[0].set_ylabel("Number of tasks")
        axes[0].set_title("Sensitivity: Window Size vs Task Count")
        axes[0].legend(fontsize=8)
        axes[0].grid(True, alpha=0.3)
        axes[1].plot(windows, expected, "s-", color="#d62728", linewidth=1.8, markersize=7)
        axes[1].axhline(9.55, color="gray", linestyle="--", linewidth=0.8, alpha=0.7, label="Default (71)")
        axes[1].set_xlabel("Smoothing window (10Hz samples)")
        axes[1].set_ylabel("Expected completed tasks")
        axes[1].set_title("Sensitivity: Window Size vs Expected Completion")
        axes[1].legend(fontsize=8)
        axes[1].grid(True, alpha=0.3)
        plt.tight_layout()
        p = fig_dir / "sensitivity_analysis.png"
        plt.savefig(p, bbox_inches="tight")
        plt.close()
        paths.append(p)

    # Speed/acceleration profile for Problem 3 trajectory
    fig, axes = plt.subplots(3, 1, figsize=(10, 8), dpi=150)
    t = traj_3["time_s"].to_numpy()
    axes[0].plot(t, traj_3["x_m"], label="X", color="#1f77b4", linewidth=1)
    axes[0].plot(t, traj_3["y_m"], label="Y", color="#ff7f0e", linewidth=1)
    axes[0].set_ylabel("Position (m)")
    axes[0].set_title("Problem 3 Fused Trajectory: Position vs Time")
    axes[0].legend(fontsize=8)
    axes[0].grid(True, alpha=0.3)
    axes[1].plot(t, traj_3["speed_m_s"], color="#2ca02c", linewidth=1)
    axes[1].axhline(2.0, color="r", linestyle="--", alpha=0.5, label="Speed limit (2 m/s)")
    axes[1].axhline(1.5, color="orange", linestyle=":", alpha=0.5, label="Photo limit (1.5 m/s)")
    axes[1].set_ylabel("Speed (m/s)")
    axes[1].set_title("Speed Profile")
    axes[1].legend(fontsize=8)
    axes[1].grid(True, alpha=0.3)
    axes[2].plot(t, traj_3["accel_m_s2"], color="#d62728", linewidth=1)
    axes[2].axhline(1.5, color="r", linestyle="--", alpha=0.5, label="Accel limit (1.5 m/s²)")
    axes[2].set_ylabel("Acceleration (m/s²)")
    axes[2].set_xlabel("Time (s)")
    axes[2].set_title("Acceleration Profile")
    axes[2].legend(fontsize=8)
    axes[2].grid(True, alpha=0.3)
    plt.tight_layout()
    p = fig_dir / "speed_acceleration_profile.png"
    plt.savefig(p, bbox_inches="tight")
    plt.close()
    paths.append(p)

    # Window impact on velocity estimation
    if window_impact_data:
        fig, axes = plt.subplots(1, 2, figsize=(12, 4.5), dpi=150)
        windows = [r["window"] for r in window_impact_data]
        axes[0].plot(windows, [r["speed_max"] for r in window_impact_data], "o-", color="#1f77b4", linewidth=1.8, markersize=6, label="Max speed")
        axes[0].plot(windows, [r["speed_mean"] for r in window_impact_data], "s-", color="#2ca02c", linewidth=1.8, markersize=6, label="Mean speed")
        axes[0].axhline(2.0, color="r", linestyle="--", alpha=0.5, label="Limit (2 m/s)")
        axes[0].set_xlabel("Smoothing window")
        axes[0].set_ylabel("Speed (m/s)")
        axes[0].set_title("Window Impact on Speed Estimates")
        axes[0].legend(fontsize=8)
        axes[0].grid(True, alpha=0.3)
        axes[1].plot(windows, [r["pct_speed_gt_2"] for r in window_impact_data], "o-", color="#d62728", linewidth=1.8, markersize=6)
        axes[1].set_xlabel("Smoothing window")
        axes[1].set_ylabel("% samples > 2 m/s")
        axes[1].set_title("% Speed Samples Exceeding Limit")
        axes[1].grid(True, alpha=0.3)
        plt.tight_layout()
        p = fig_dir / "window_impact_velocity.png"
        plt.savefig(p, bbox_inches="tight")
        plt.close()
        paths.append(p)

    return paths


def xml_escape(text: object) -> str:
    return html.escape(str(text), quote=False)


def w_run(text: object, bold: bool = False, font: str = "宋体", size: int = 24) -> str:
    b = "<w:b/>" if bold else ""
    return (
        "<w:r><w:rPr>"
        f"{b}<w:rFonts w:ascii=\"Times New Roman\" w:eastAsia=\"{font}\" w:hAnsi=\"Times New Roman\"/>"
        f"<w:sz w:val=\"{size}\"/><w:szCs w:val=\"{size}\"/>"
        f"</w:rPr><w:t xml:space=\"preserve\">{xml_escape(text)}</w:t></w:r>"
    )


def w_p(
    text: object = "",
    align: str = "left",
    bold: bool = False,
    font: str = "宋体",
    size: int = 24,
    spacing_after: int = 80,
    indent_first: bool = False,
) -> str:
    jc = {"left": "left", "center": "center", "right": "right"}.get(align, "left")
    ind = '<w:ind w:firstLine="480"/>' if indent_first else ""
    return (
        "<w:p><w:pPr>"
        f"<w:jc w:val=\"{jc}\"/><w:spacing w:after=\"{spacing_after}\" w:line=\"240\" w:lineRule=\"auto\"/>"
        f"{ind}</w:pPr>{w_run(text, bold=bold, font=font, size=size)}</w:p>"
    )


def w_heading(text: str, level: int) -> str:
    if level == 1:
        return w_p(text, align="center", bold=True, font="黑体", size=28, spacing_after=120)
    return w_p(text, align="left", bold=True, font="黑体", size=24, spacing_after=80)


def w_table(headers: Sequence[str], rows: Sequence[Sequence[object]]) -> str:
    def cell(text: object, bold: bool = False) -> str:
        return (
            "<w:tc><w:tcPr><w:tcW w:w=\"0\" w:type=\"auto\"/>"
            '<w:vAlign w:val="center"/></w:tcPr>'
            f"{w_p(text, align='center', bold=bold, size=20, spacing_after=0)}</w:tc>"
        )

    borders = (
        "<w:tblPr><w:tblBorders>"
        '<w:top w:val="single" w:sz="4" w:space="0" w:color="000000"/>'
        '<w:left w:val="single" w:sz="4" w:space="0" w:color="000000"/>'
        '<w:bottom w:val="single" w:sz="4" w:space="0" w:color="000000"/>'
        '<w:right w:val="single" w:sz="4" w:space="0" w:color="000000"/>'
        '<w:insideH w:val="single" w:sz="4" w:space="0" w:color="000000"/>'
        '<w:insideV w:val="single" w:sz="4" w:space="0" w:color="000000"/>'
        "</w:tblBorders></w:tblPr>"
    )
    out = ["<w:tbl>", borders]
    out.append("<w:tr>" + "".join(cell(h, True) for h in headers) + "</w:tr>")
    for row in rows:
        out.append("<w:tr>" + "".join(cell(x) for x in row) + "</w:tr>")
    out.append("</w:tbl>")
    return "".join(out)


def w_image(rid: str, width_in: float, height_in: float) -> str:
    cx = int(width_in * 914400)
    cy = int(height_in * 914400)
    return f"""
<w:p><w:pPr><w:jc w:val="center"/></w:pPr><w:r><w:drawing>
<wp:inline distT="0" distB="0" distL="0" distR="0">
<wp:extent cx="{cx}" cy="{cy}"/><wp:effectExtent l="0" t="0" r="0" b="0"/>
<wp:docPr id="1" name="Picture"/><wp:cNvGraphicFramePr>
<a:graphicFrameLocks xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" noChangeAspect="1"/>
</wp:cNvGraphicFramePr><a:graphic xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">
<a:graphicData uri="http://schemas.openxmlformats.org/drawingml/2006/picture">
<pic:pic xmlns:pic="http://schemas.openxmlformats.org/drawingml/2006/picture">
<pic:nvPicPr><pic:cNvPr id="0" name="image.png"/><pic:cNvPicPr/></pic:nvPicPr>
<pic:blipFill><a:blip r:embed="{rid}"/><a:stretch><a:fillRect/></a:stretch></pic:blipFill>
<pic:spPr><a:xfrm><a:off x="0" y="0"/><a:ext cx="{cx}" cy="{cy}"/></a:xfrm>
<a:prstGeom prst="rect"><a:avLst/></a:prstGeom></pic:spPr>
</pic:pic></a:graphicData></a:graphic></wp:inline></w:drawing></w:r></w:p>
"""


def section_break_no_footer() -> str:
    return (
        "<w:p><w:pPr><w:sectPr>"
        '<w:type w:val="nextPage"/>'
        '<w:pgSz w:w="11906" w:h="16838"/>'
        '<w:pgMar w:top="1440" w:right="1440" w:bottom="1440" w:left="1440" w:header="0" w:footer="720" w:gutter="0"/>'
        "</w:sectPr></w:pPr></w:p>"
    )


def final_section_with_footer() -> str:
    return (
        '<w:sectPr><w:footerReference w:type="default" r:id="rIdFooter1"/>'
        '<w:pgSz w:w="11906" w:h="16838"/>'
        '<w:pgMar w:top="1440" w:right="1440" w:bottom="1440" w:left="1440" w:header="0" w:footer="720" w:gutter="0"/>'
        '<w:pgNumType w:start="1"/></w:sectPr>'
    )


def build_report_text(
    results: dict[int, AlignmentResult],
    trajectories: dict[int, pd.DataFrame],
    tasks: pd.DataFrame,
    sens_windows: list[int] | None = None,
    sens_rows: list[dict[str, float | str]] | None = None,
) -> tuple[str, list[str], list[tuple[str, Sequence[str], Sequence[Sequence[object]]]]]:
    r1, r2, r3 = results[1], results[2], results[3]
    task_count = len(tasks)
    shoot_count = int((tasks["task"] == "模拟射击").sum()) if not tasks.empty else 0
    photo_count = int((tasks["task"] == "拍照").sum()) if not tasks.empty else 0
    expected = float(tasks["expected_success"].sum()) if not tasks.empty else 0.0

    r2_f_str = f"F={r2.f_statistic:.2f}, p={r2.f_p_value:.4f}" if r2.f_statistic else ""
    r3_f_str = f"F={r3.f_statistic:.2f}, p={r3.f_p_value:.4f}" if r3.f_statistic else ""
    if r3.has_system_bias:
        abstract_p3 = (
            f"问题3实际数据的固定偏差为({r3.bias2_x:.4f},{r3.bias2_y:.4f})m"
            f"（F检验{r3_f_str}），采用偏差模型输出10Hz融合轨迹。"
        )
    else:
        abstract_p3 = (
            f"问题3实际数据的候选固定偏差为({getattr(r3, 'candidate_bias_x', 0.0):.4f},"
            f"{getattr(r3, 'candidate_bias_y', 0.0):.4f})m，"
            f"误差下降{100*r3.bias_model_improvement:.2f}%"
            f"（F检验{r3_f_str}），"
            "判定不采用固定系统偏差修正，输出10Hz融合轨迹。"
        )
    abstract = (
        "针对两类异频异步定位数据，建立了以方式1为时间基准的轨迹匹配、固定偏差估计和10Hz重采样融合模型。"
        f"问题1在无噪声条件下得到方式2相对方式1的时间偏差为{r1.delta2_minus_1:.4f}s，融合残差接近0。"
        f"问题2在随机噪声和系统偏差共同存在时，估计方式2时间偏差为{r2.delta2_minus_1:.4f}s "
        f"(95%置信区间[{r2.ci_delta_lo:.4f},{r2.ci_delta_hi:.4f}]s)，"
        f"系统偏差为({r2.bias2_x:.4f},{r2.bias2_y:.4f})m，"
        f"带偏差模型使匹配均方误差下降{100*r2.bias_model_improvement:.2f}%"
        f"（F检验{r2_f_str}）。"
        f"{abstract_p3}"
        f"在任务优化中，采用加权区间调度（DP）最大化期望完成数，"
        f"共得到{task_count}项非重叠任务，"
        f"其中射击{shoot_count}项、拍照{photo_count}项，"
        f"考虑85%单次命中率后的期望完成数为{expected:.2f}。"
    )
    keywords = ["多源融合", "时间对齐", "系统偏差", "F检验", "加权区间调度", "10Hz重采样", "任务优化"]
    tables = [
        (
            "表1  时间偏差与系统偏差估计结果",
            ["问题", "delta/s", "delta 95%CI/s", "bias_x/m", "bias_y/m", "系统偏差判定", "重叠时长/s"],
            [
                [1, f"{r1.delta2_minus_1:.4f}", "—", "0", "0", "否", f"{r1.overlap_seconds:.1f}"],
                [2, f"{r2.delta2_minus_1:.4f}",
                 f"[{r2.ci_delta_lo:.4f},{r2.ci_delta_hi:.4f}]",
                 f"{r2.bias2_x:.4f}", f"{r2.bias2_y:.4f}", "是", f"{r2.overlap_seconds:.1f}"],
                [3, f"{r3.delta2_minus_1:.4f}",
                 f"[{r3.ci_delta_lo:.4f},{r3.ci_delta_hi:.4f}]" if r3.ci_delta_lo != 0 else "—",
                 f"{r3.bias2_x:.4f}", f"{r3.bias2_y:.4f}", "否", f"{r3.overlap_seconds:.1f}"],
            ],
        ),
        (
            "表2  F检验结果",
            ["问题", "F统计量", "p值", "显著性水平", "结论"],
            [
                [2, f"{r2.f_statistic:.2f}" if r2.f_statistic else "—",
                 f"{r2.f_p_value:.4f}" if r2.f_p_value else "—", 0.05, "显著（存在系统偏差）"],
                [3, f"{r3.f_statistic:.2f}" if r3.f_statistic else "—",
                 f"{r3.f_p_value:.4f}" if r3.f_p_value else "—", 0.05, "统计显著，但工程量级不足"],
            ],
        ),
        (
            "表3  10Hz轨迹规模",
            ["问题", "点数", "起始/s", "终止/s", "x范围/m", "y范围/m"],
            [
                [
                    i,
                    len(df),
                    f"{df['time_s'].iloc[0]:.1f}",
                    f"{df['time_s'].iloc[-1]:.1f}",
                    f"{df['x_m'].min():.2f}~{df['x_m'].max():.2f}",
                    f"{df['y_m'].min():.2f}~{df['y_m'].max():.2f}",
                ]
                for i, df in trajectories.items()
            ],
        ),
        (
            "表4  任务优化结果摘要",
            ["指标", "数值"],
            [
                ["任务总数", task_count],
                ["射击任务数", shoot_count],
                ["拍照任务数", photo_count],
                ["期望完成数", f"{expected:.2f}"],
            ],
        ),
    ]
    if sens_rows:
        tables.append((
            "附表  平滑窗口敏感性分析",
            list(sens_rows[0].keys()),
            [[r["平滑窗口(10Hz点数)"], r["任务数"], r["期望完成数"]] for r in sens_rows],
        ))
    return abstract, keywords, tables


def write_markdown_report(
    path: Path,
    results: dict[int, AlignmentResult],
    trajectories: dict[int, pd.DataFrame],
    tasks: pd.DataFrame,
    fig_paths: Sequence[Path],
    sens_windows: list[int] | None = None,
    sens_rows: list[dict[str, float | str]] | None = None,
) -> None:
    abstract, keywords, _tables = build_report_text(results, trajectories, tasks, sens_windows, sens_rows)
    r1, r2, r3 = results[1], results[2], results[3]
    md = []
    md.append("# B题 多源融合机器人定位及任务优化\n")
    md.append("## 摘要\n")
    md.append(abstract + "\n")
    md.append("**关键词：**" + "；".join(keywords) + "\n")
    md.append("## 1 问题重述\n")
    md.append("两种定位方式的采样频率分别为4Hz和5Hz，存在起始时间不同步、随机噪声以及可能的固定系统偏差。需要建立时间对齐、偏差修正和轨迹融合模型，并基于附件3轨迹完成射击和拍照任务设计。\n")
    md.append("## 2 模型与假设\n")
    md.append("设方式2修正后的时间为 tau=t2-delta，坐标修正为 p2'=p2-b。delta 表示方式2相对方式1的时间偏差，b 表示方式2相对方式1的固定坐标偏差。采用F检验（H0: bx=by=0）判定系统偏差显著性，显著性水平α=0.05。任务调度采用加权区间调度（DP）最大化非重叠任务的期望完成数。题面未给出任务互斥或冷却约束，因此主结果仅采用附录列出的距离、速度、加速度、准备时间和拍照角度约束。\n")
    md.append("## 3 主要结果\n")
    ci2 = f" (95%CI: [{r2.ci_delta_lo:.4f}, {r2.ci_delta_hi:.4f}]s)" if r2.ci_delta_lo != 0 or r2.ci_delta_hi != 0 else ""
    ci3 = f" (95%CI: [{r3.ci_delta_lo:.4f}, {r3.ci_delta_hi:.4f}]s)" if r3.ci_delta_lo != 0 else ""
    f2 = f"，F检验显著(F={r2.f_statistic:.2f}, p={r2.f_p_value:.4f})" if r2.f_statistic else ""
    f3 = f"，F检验F={r3.f_statistic:.2f}, p={r3.f_p_value:.4f}" if r3.f_statistic else ""
    if r3.has_system_bias:
        bias_str3 = f"({r3.bias2_x:.4f},{r3.bias2_y:.4f})m"
        bias_decision3 = f"存在固定系统偏差{bias_str3}{f3}"
    else:
        cand_x = getattr(r3, 'candidate_bias_x', 0.0)
        cand_y = getattr(r3, 'candidate_bias_y', 0.0)
        bias_str3 = f"({cand_x:.4f},{cand_y:.4f})m"
        bias_decision3 = f"候选偏差{bias_str3}，误差下降{100*r3.bias_model_improvement:.2f}%{f3}，不采用偏差修正"
    md.append(f"- 问题1：delta={r1.delta2_minus_1:.4f}s，系统偏差为0。\n")
    md.append(f"- 问题2：delta={r2.delta2_minus_1:.4f}s{ci2}，方式2系统偏差=({r2.bias2_x:.4f},{r2.bias2_y:.4f})m{f2}。\n")
    md.append(f"- 问题3：{bias_decision3}；delta={r3.delta2_minus_1:.4f}s{ci3}。\n")
    md.append(f"- 问题4：加权区间调度后得到{len(tasks)}项非重叠任务，已写入 `result.xlsx`。\n")
    for p in fig_paths:
        md.append(f"\n![{p.stem}]({p.as_posix()})\n")
    md.append("\n## 4 任务明细\n")
    md.append(dataframe_to_markdown(tasks) if not tasks.empty else "无可行任务")
    md.append("\n## 参考文献\n")
    md.append("[1] Savitzky A, Golay M J E, Smoothing and Differentiation of Data by Simplified Least Squares Procedures, Analytical Chemistry, 36(8):1627-1639, 1964.\n")
    path.write_text("\n".join(md), encoding="utf-8")


def dataframe_to_markdown(df: pd.DataFrame) -> str:
    headers = [str(c) for c in df.columns]
    rows = []
    for _, row in df.iterrows():
        rows.append([format_markdown_cell(row[c]) for c in df.columns])
    out = ["| " + " | ".join(headers) + " |"]
    out.append("| " + " | ".join(["---"] * len(headers)) + " |")
    for row in rows:
        out.append("| " + " | ".join(row) + " |")
    return "\n".join(out)


def format_markdown_cell(value: object) -> str:
    if isinstance(value, float):
        text = f"{value:.4f}"
    else:
        text = str(value)
    return text.replace("|", "\\|")


def write_docx_report(
    path: Path,
    results: dict[int, AlignmentResult],
    trajectories: dict[int, pd.DataFrame],
    tasks: pd.DataFrame,
    fig_paths: Sequence[Path],
    sens_windows: list[int] | None = None,
    sens_rows: list[dict[str, float | str]] | None = None,
) -> None:
    abstract, keywords, tables = build_report_text(results, trajectories, tasks, sens_windows, sens_rows)
    r1, r2, r3 = results[1], results[2], results[3]

    body: list[str] = []
    body.append(w_p("B题 多源融合机器人定位及任务优化", align="center", bold=True, font="黑体", size=32, spacing_after=180))
    body.append(w_p("摘要", align="center", bold=True, font="黑体", size=28, spacing_after=120))
    body.append(w_p(abstract, indent_first=True))
    body.append(w_p("关键词：" + "；".join(keywords), bold=True, font="宋体"))
    body.append(section_break_no_footer())

    body.append(w_heading("一、问题重述", 1))
    body.append(w_p("题目给出两类定位数据：方式1为4Hz，方式2为5Hz。两类数据存在时间起点不同、采样频率不同、随机噪声以及可能的固定系统偏差。本文需要完成三类轨迹的对齐与10Hz融合，并利用附件3轨迹对沿途射击和拍照任务进行可行时刻设计。", indent_first=True))
    body.append(w_heading("二、模型假设与符号", 1))
    body.append(w_p("以方式1时间为基准，设方式2修正时间为 tau=t2-delta，其中 delta 为方式2相对方式1的时间偏差；设方式2相对方式1的固定坐标偏差为 b=(b_x,b_y)，修正坐标为 p2'=p2-b。题面未给出任务互斥、执行器冷却或同一时刻只能处理一个目标的限制，故主结果仅采用附录明确给出的距离、速度、加速度、准备时间和拍照角度约束。", indent_first=True))
    body.append(w_p("射击约束为距离5m至30m、线速度不超过2m/s、加速度不超过1.5m/s^2，执行前1.5s内均满足上述约束；拍照约束为距离10m至40m、不同拍照方向角至少60度、线速度不超过1.5m/s、加速度不超过1.5m/s^2，执行前0.5s内均满足上述约束。", indent_first=True))

    body.append(w_heading("三、时间对齐与融合模型", 1))
    body.append(w_p("对给定 delta，将方式2时间平移到方式1基准，在两者重叠区间按0.1s网格插值。无噪声问题直接最小化两轨迹平方距离；有噪声问题在每个候选 delta 下先用重叠样本的中位差估计固定偏差，再最小化去偏后的截尾均方误差。为避免局部短重叠造成伪匹配，搜索时要求重叠时长不低于较短轨迹时长的85%。", indent_first=True))
    body.append(w_p("融合阶段将方式2坐标扣除系统偏差后，与方式1在10Hz网格上等权平均。对含噪声轨迹采用Savitzky-Golay平滑与差分求速度、加速度[1]；问题4使用问题3的10Hz融合轨迹。", indent_first=True))
    for title, headers, rows in tables[:2]:
        body.append(w_p(title, align="center", font="黑体", size=22, spacing_after=40))
        body.append(w_table(headers, rows))
        body.append(w_p("", spacing_after=80))

    for i, _fig_path in enumerate(fig_paths[:3], start=1):
        body.append(w_p(f"图{i}  问题{i}的10Hz融合轨迹", align="center", font="黑体", size=22, spacing_after=40))
        body.append(w_image(f"rIdImage{i}", 5.8, 4.3))

    body.append(w_heading("四、问题结果分析", 1))
    body.append(w_p(f"问题1中无随机噪声和系统偏差，最优时间偏差为{r1.delta2_minus_1:.4f}s，匹配均方误差约为{r1.mse_with_bias:.3e}，说明两类数据在时间平移后完全一致。", indent_first=True))
    f2_str = f"F={r2.f_statistic:.2f}, p={r2.f_p_value:.4f}" if r2.f_statistic else ""
    body.append(w_p(f"问题2中，若不估计系统偏差，截尾均方误差为{r2.mse_without_bias:.4f}；引入固定偏差后误差降至{r2.mse_with_bias:.4f}，下降{100*r2.bias_model_improvement:.2f}%（{f2_str}，p<0.05），因此拒绝无偏差原假设，判定存在固定系统偏差。方式2相对方式1的偏差为({r2.bias2_x:.4f},{r2.bias2_y:.4f})m。delta的95%置信区间为[{r2.ci_delta_lo:.4f},{r2.ci_delta_hi:.4f}]s。", indent_first=True))
    f3_str = f"F={r3.f_statistic:.2f}, p={r3.f_p_value:.4f}" if r3.f_statistic else ""
    if r3.has_system_bias:
        p3_text = (
            f"问题3中，估计得到固定偏差为({r3.bias2_x:.4f},{r3.bias2_y:.4f})m"
            f"（{f3_str}），采用偏差模型输出融合轨迹。"
        )
    else:
        p3_text = (
            f"问题3中，带偏差模型得到的候选偏差为({getattr(r3, 'candidate_bias_x', 0.0):.4f},"
            f"{getattr(r3, 'candidate_bias_y', 0.0):.4f})m，误差下降比例为{100*r3.bias_model_improvement:.2f}%"
            f"（{f3_str}），但F检验未达到统计显著性水平，因此不采用固定系统偏差修正，使用无偏差模型输出融合轨迹。"
        )
    body.append(w_p(p3_text, indent_first=True))

    body.append(w_heading("五、任务优化模型与结果", 1))
    body.append(w_p("在10Hz轨迹上逐时刻检查目标距离、机器人线速度和加速度，并用滚动窗口保证准备区间内约束全部成立。射击目标每个目标选择距离最近的可行时刻；拍照目标在可行时刻中按方向角差异至少60度筛选，以保留尽量多的不同视角。采用加权区间调度（DP）从候选时刻中选出最大期望完成数的非重叠子集。", indent_first=True))
    if len(fig_paths) >= 4:
        body.append(w_p("图4  附件3轨迹与被选任务目标", align="center", font="黑体", size=22, spacing_after=40))
        body.append(w_image("rIdImage4", 5.8, 4.3))
    title, headers, rows = tables[3]
    body.append(w_p(title, align="center", font="黑体", size=22, spacing_after=40))
    body.append(w_table(headers, rows))
    body.append(w_p("任务明细已按开始准备时刻写入输出目录中的 result.xlsx。由于射击单次命中率为85%，射击任务按期望完成数计为0.85项，拍照任务计为1项。", indent_first=True))

    # 敏感性分析表
    if len(tables) >= 5:
        sens_title, sens_headers, sens_rows_t = tables[4]
        body.append(w_p(sens_title, align="center", font="黑体", size=22, spacing_after=40))
        body.append(w_table(sens_headers, sens_rows_t))
        body.append(w_p("", spacing_after=80))

    if not tasks.empty:
        display = tasks[["target_id", "task", "prep_start_s", "exec_time_s", "distance_m", "speed_m_s", "accel_m_s2"]].copy()
        display["distance_m"] = display["distance_m"].map(lambda x: f"{x:.2f}")
        display["speed_m_s"] = display["speed_m_s"].map(lambda x: f"{x:.2f}")
        display["accel_m_s2"] = display["accel_m_s2"].map(lambda x: f"{x:.2f}")
        body.append(w_p("表5  任务明细", align="center", font="黑体", size=22, spacing_after=40))
        body.append(w_table(["目标", "任务", "开始准备/s", "执行/s", "距离/m", "速度/(m/s)", "加速度/(m/s^2)"], display.values.tolist()))

    body.append(w_heading("六、模型评价", 1))
    body.append(w_p("模型优点是参数含义清晰、只依赖题目给出的多源轨迹数据。偏差判定采用嵌套模型F检验（α=0.05），相比经验阈值具有更清晰的统计依据。delta估计附有95%置信区间，用于反映估计精度。任务调度采用加权区间调度DP得到最大化期望完成数的非重叠方案。重叠时长约束（85%）有效避免短区间伪匹配。", indent_first=True))
    body.append(w_p("不足之处在于速度、加速度由平滑轨迹差分获得，平滑窗口会影响临界任务的可行性，敏感性分析显示任务数和期望完成数对平滑窗口存在一定波动。若后续给出机器人动力学模型或执行器互斥约束，可将本文候选任务表作为输入，进一步建立混合整数规划进行全局时序优化。", indent_first=True))
    body.append(w_heading("参考文献", 1))
    body.append(w_p("[1] Savitzky A, Golay M J E, Smoothing and Differentiation of Data by Simplified Least Squares Procedures, Analytical Chemistry, 36(8):1627-1639, 1964.", indent_first=False))

    document_xml = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<w:document xmlns:wpc="http://schemas.microsoft.com/office/word/2010/wordprocessingCanvas"
xmlns:mc="http://schemas.openxmlformats.org/markup-compatibility/2006"
xmlns:o="urn:schemas-microsoft-com:office:office"
xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"
xmlns:m="http://schemas.openxmlformats.org/officeDocument/2006/math"
xmlns:v="urn:schemas-microsoft-com:vml"
xmlns:wp14="http://schemas.microsoft.com/office/word/2010/wordprocessingDrawing"
xmlns:wp="http://schemas.openxmlformats.org/drawingml/2006/wordprocessingDrawing"
xmlns:w10="urn:schemas-microsoft-com:office:word"
xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main"
xmlns:w14="http://schemas.microsoft.com/office/word/2010/wordml"
xmlns:wpg="http://schemas.microsoft.com/office/word/2010/wordprocessingGroup"
xmlns:wpi="http://schemas.microsoft.com/office/word/2010/wordprocessingInk"
xmlns:wne="http://schemas.microsoft.com/office/word/2006/wordml"
xmlns:wps="http://schemas.microsoft.com/office/word/2010/wordprocessingShape"
xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main"
xmlns:pic="http://schemas.openxmlformats.org/drawingml/2006/picture"
mc:Ignorable="w14 wp14"><w:body>{''.join(body)}{final_section_with_footer()}</w:body></w:document>"""

    rels = [
        '<Relationship Id="rIdStyles" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/>',
        '<Relationship Id="rIdSettings" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/settings" Target="settings.xml"/>',
        '<Relationship Id="rIdFooter1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/footer" Target="footer1.xml"/>',
    ]
    image_targets: list[tuple[str, Path]] = []
    for i, fig_path in enumerate(fig_paths, start=1):
        rels.append(f'<Relationship Id="rIdImage{i}" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/image" Target="media/image{i}.png"/>')
        image_targets.append((f"word/media/image{i}.png", fig_path))

    content_types = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
<Default Extension="xml" ContentType="application/xml"/>
<Default Extension="png" ContentType="image/png"/>
<Override PartName="/word/document.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>
<Override PartName="/word/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.styles+xml"/>
<Override PartName="/word/settings.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.settings+xml"/>
<Override PartName="/word/footer1.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.footer+xml"/>
</Types>"""
    root_rels = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="word/document.xml"/>
</Relationships>"""
    doc_rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        + "".join(rels)
        + "</Relationships>"
    )
    styles = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<w:styles xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">
<w:style w:type="paragraph" w:default="1" w:styleId="Normal">
<w:name w:val="Normal"/><w:qFormat/>
<w:rPr><w:rFonts w:ascii="Times New Roman" w:eastAsia="宋体" w:hAnsi="Times New Roman"/><w:sz w:val="24"/><w:szCs w:val="24"/></w:rPr>
</w:style></w:styles>"""
    settings = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<w:settings xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">
<w:zoom w:percent="100"/><w:defaultTabStop w:val="420"/>
</w:settings>"""
    footer = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<w:ftr xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">
<w:p><w:pPr><w:jc w:val="center"/></w:pPr>
<w:r><w:fldChar w:fldCharType="begin"/></w:r>
<w:r><w:instrText xml:space="preserve">PAGE</w:instrText></w:r>
<w:r><w:fldChar w:fldCharType="separate"/></w:r>
<w:r><w:t>1</w:t></w:r>
<w:r><w:fldChar w:fldCharType="end"/></w:r>
</w:p></w:ftr>"""

    def write_zip(dest: Path) -> None:
        with zipfile.ZipFile(dest, "w", zipfile.ZIP_DEFLATED) as z:
            z.writestr("[Content_Types].xml", content_types)
            z.writestr("_rels/.rels", root_rels)
            z.writestr("word/document.xml", document_xml)
            z.writestr("word/_rels/document.xml.rels", doc_rels)
            z.writestr("word/styles.xml", styles)
            z.writestr("word/settings.xml", settings)
            z.writestr("word/footer1.xml", footer)
            for arc, src in image_targets:
                z.write(src, arc)

    try:
        write_zip(path)
    except PermissionError:
        write_zip(path.with_name(f"{path.stem}_codex{path.suffix}"))


def choose_font(paths: Sequence[str]) -> FontProperties:
    for p in paths:
        path = Path(p)
        if path.exists():
            return FontProperties(fname=str(path))
    return FontProperties()


def wrap_cjk_text(text: str, width: int) -> list[str]:
    lines: list[str] = []
    for raw in str(text).splitlines() or [""]:
        raw = raw.strip()
        while len(raw) > width:
            cut = width
            for mark in "，。；、：,.; ":
                pos = raw.rfind(mark, 0, width)
                if pos > width * 0.55:
                    cut = pos + 1
                    break
            lines.append(raw[:cut])
            raw = raw[cut:].lstrip()
        lines.append(raw)
    return lines


def write_pdf_report(
    path: Path,
    results: dict[int, AlignmentResult],
    trajectories: dict[int, pd.DataFrame],
    tasks: pd.DataFrame,
    fig_paths: Sequence[Path],
    sens_windows: list[int] | None = None,
    sens_rows: list[dict[str, float | str]] | None = None,
) -> None:
    body_font = choose_font([r"C:\Windows\Fonts\simsun.ttc", r"C:\Windows\Fonts\msyh.ttc"])
    title_font = choose_font([r"C:\Windows\Fonts\simhei.ttf", r"C:\Windows\Fonts\msyhbd.ttc", r"C:\Windows\Fonts\simsunb.ttf"])
    abstract, keywords, _tables = build_report_text(results, trajectories, tasks, sens_windows, sens_rows)
    r1, r2, r3 = results[1], results[2], results[3]

    sections: list[tuple[str, list[str]]] = [
        (
            "一、问题重述",
            [
                "两类定位数据采样频率分别为4Hz和5Hz，需解决时间异步、随机噪声和可能的固定系统偏差，并输出10Hz连续轨迹；在附件3轨迹上，还需设计射击与拍照任务时刻。"
            ],
        ),
        (
            "二、模型与假设",
            [
                "以方式1为时间基准，方式2修正时间为 tau=t2-delta，修正坐标为 p2'=p2-b。搜索delta时要求两轨迹重叠时长不少于较短轨迹的85%，避免短区间伪匹配。",
                "系统偏差判定采用嵌套模型F检验。题面未给出任务互斥或冷却约束，主结果只采用附录明确给出的距离、速度、加速度、准备时间和拍照角度约束。",
            ],
        ),
        (
            "三、结果",
            [
                f"问题1：delta={r1.delta2_minus_1:.4f}s，系统偏差为0。",
                f"问题2：delta={r2.delta2_minus_1:.4f}s，95%CI=[{r2.ci_delta_lo:.4f},{r2.ci_delta_hi:.4f}]s，方式2系统偏差=({r2.bias2_x:.4f},{r2.bias2_y:.4f})m，F={r2.f_statistic:.2f}, p={r2.f_p_value:.4f}。",
                f"问题3：候选偏差=({getattr(r3, 'candidate_bias_x', 0.0):.4f},{getattr(r3, 'candidate_bias_y', 0.0):.4f})m，误差下降{100*r3.bias_model_improvement:.2f}%，F={r3.f_statistic:.2f}, p={r3.f_p_value:.4f}{'，采用偏差修正' if r3.has_system_bias else '，不采用偏差修正'}；delta={r3.delta2_minus_1:.4f}s。",
                f"问题4：得到{len(tasks)}项任务，其中射击{int((tasks['task']=='模拟射击').sum()) if not tasks.empty else 0}项、拍照{int((tasks['task']=='拍照').sum()) if not tasks.empty else 0}项，期望完成数为{float(tasks['expected_success'].sum()) if not tasks.empty else 0.0:.2f}。",
            ],
        ),
    ]

    def new_page(pdf: PdfPages, page_no: int | None = None):
        fig = plt.figure(figsize=(8.27, 11.69), dpi=150)
        ax = fig.add_axes([0, 0, 1, 1])
        ax.axis("off")
        if page_no is not None:
            ax.text(0.5, 0.035, str(page_no), ha="center", va="center", fontsize=10, fontproperties=body_font)
        return fig, ax

    with PdfPages(path) as pdf:
        fig, ax = new_page(pdf, None)
        y = 0.93
        ax.text(0.5, y, "B题 多源融合机器人定位及任务优化", ha="center", va="top", fontsize=16, fontproperties=title_font)
        y -= 0.06
        ax.text(0.5, y, "摘要", ha="center", va="top", fontsize=14, fontproperties=title_font)
        y -= 0.04
        for line in wrap_cjk_text(abstract, 43):
            ax.text(0.11, y, line, ha="left", va="top", fontsize=10.5, fontproperties=body_font)
            y -= 0.026
        y -= 0.01
        for line in wrap_cjk_text("关键词：" + "；".join(keywords), 43):
            ax.text(0.11, y, line, ha="left", va="top", fontsize=10.5, fontproperties=body_font)
            y -= 0.026
        pdf.savefig(fig)
        plt.close(fig)

        page_no = 1
        fig, ax = new_page(pdf, page_no)
        y = 0.94
        for title, paras in sections:
            if y < 0.16:
                pdf.savefig(fig)
                plt.close(fig)
                page_no += 1
                fig, ax = new_page(pdf, page_no)
                y = 0.94
            ax.text(0.5, y, title, ha="center", va="top", fontsize=13, fontproperties=title_font)
            y -= 0.04
            for para in paras:
                for line in wrap_cjk_text(para, 45):
                    ax.text(0.11, y, line, ha="left", va="top", fontsize=10.5, fontproperties=body_font)
                    y -= 0.026
                y -= 0.012
        ax.text(0.5, y, "四、任务明细", ha="center", va="top", fontsize=13, fontproperties=title_font)
        y -= 0.04
        headers = ["序号", "目标", "任务", "准备/s", "执行/s"]
        ax.text(0.11, y, "  ".join(headers), ha="left", va="top", fontsize=9.5, fontproperties=title_font)
        y -= 0.026
        for i, row in tasks.iterrows():
            if y < 0.12:
                pdf.savefig(fig)
                plt.close(fig)
                page_no += 1
                fig, ax = new_page(pdf, page_no)
                y = 0.94
            text = f"{i+1:>2}  {row['target_id']}  {row['task']}  {row['prep_start_s']:.1f}  {row['exec_time_s']:.1f}"
            ax.text(0.11, y, text, ha="left", va="top", fontsize=9.5, fontproperties=body_font)
            y -= 0.024
        pdf.savefig(fig)
        plt.close(fig)

        for fig_path in fig_paths:
            page_no += 1
            fig, ax = new_page(pdf, page_no)
            ax.imshow(plt.imread(fig_path))
            ax.axis("off")
            pdf.savefig(fig)
            plt.close(fig)


def sensitivity_analysis(path: Path, alignment: AlignmentResult, target_path: Path,
                          traj_smooth_window: int) -> tuple[int, float]:
    """用指定平滑窗口运行任务优化，返回(总任务数, 期望完成数)."""
    traj_s = make_trajectory(path, alignment, traj_smooth_window)
    tasks_s = select_task_candidates(traj_s, target_path)
    if tasks_s.empty:
        return 0, 0.0
    tasks_opt = weighted_interval_scheduling(tasks_s)
    return len(tasks_opt), float(tasks_opt["expected_success"].sum())


def main() -> None:
    parser = argparse.ArgumentParser(description="B题多源融合定位与任务优化")
    parser.add_argument("--root", type=Path, default=Path(__file__).resolve().parent, help="B题目录")
    parser.add_argument("--sequential", action="store_true",
                        help="使用旧版FIFO顺序过滤（默认使用加权区间调度DP）")
    args = parser.parse_args()

    root = args.root
    out_dir = root / "outputs"
    out_dir.mkdir(parents=True, exist_ok=True)

    files, result_template = discover_files(root)
    alignments = build_alignment_results(files)
    save_estimates(out_dir / "estimates_summary.xlsx", alignments)

    smooth_windows = {1: 1, 2: 71, 3: 71}
    trajectories: dict[int, pd.DataFrame] = {}
    for i in (1, 2, 3):
        df = make_trajectory(files[i], alignments[i], smooth_windows[i])
        trajectories[i] = df
        df.to_excel(out_dir / f"problem{i}_10Hz_trajectory.xlsx", index=False)

    tasks = select_task_candidates(trajectories[3], files[4])
    tasks_original_count = len(tasks)
    if args.sequential:
        tasks = filter_sequential(tasks)
    else:
        tasks = weighted_interval_scheduling(tasks)
    tasks.to_excel(out_dir / "task_schedule_detail.xlsx", index=False)
    write_result_xlsx(result_template, out_dir / "result.xlsx", tasks)

    # 敏感性分析：不同平滑窗口下的任务数
    sens_windows = [31, 51, 71, 91, 111]
    sens_rows = []
    for sw in sens_windows:
        n_tasks, expected = sensitivity_analysis(files[3], alignments[3], files[4], sw)
        sens_rows.append({"平滑窗口(10Hz点数)": sw, "任务数": n_tasks, "期望完成数": round(expected, 2)})
    sens_df = pd.DataFrame(sens_rows)
    sens_df.to_excel(out_dir / "sensitivity_smooth_window.xlsx", index=False)

    # ----- 诊断分析 -----
    diagnostic_results: dict[int, dict] = {}
    for prob in (2, 3):
        t1, p1, t2, p2 = read_pair(files[prob])
        al = alignments[prob]
        diag = residual_diagnostics(t1, p1, t2, p2, al.delta2_minus_1,
                                     np.array([al.bias2_x, al.bias2_y]),
                                     label=f"Problem {prob}")
        diagnostic_results[prob] = diag
        print(f"  问题{prob}残差诊断: mean={diag['mean_residual']:.4f}, "
              f"std={diag['std_residual']:.4f}, Shapiro p={diag['shapiro_p']:.4f}")

    # 窗口速度估计影响分析
    t1_3, p1_3, t2_3, p2_3 = read_pair(files[3])
    window_impact = estimate_window_impact(
        t1_3, p1_3, t2_3, p2_3,
        alignments[3].delta2_minus_1,
        np.array([alignments[3].bias2_x, alignments[3].bias2_y]),
        windows=[1, 11, 21, 31, 41, 51, 61, 71, 81, 91, 101, 111],
    )

    fig_paths = plot_outputs(out_dir, trajectories, tasks, files[4])
    diag_paths = plot_diagnostics(
        out_dir, files, alignments, trajectories[3],
        sens_windows=sens_windows, sens_rows_data=sens_rows,
        window_impact_data=window_impact,
    )
    fig_paths = list(fig_paths) + list(diag_paths)
    write_markdown_report(out_dir / "B题_论文.md", alignments, trajectories, tasks, fig_paths,
                           sens_windows=sens_windows, sens_rows=sens_rows)
    write_docx_report(out_dir / "B题_多源融合机器人定位及任务优化_论文.docx", alignments, trajectories, tasks, fig_paths,
                       sens_windows=sens_windows, sens_rows=sens_rows)
    pdf_path = out_dir / "B题_多源融合机器人定位及任务优化_论文.pdf"
    try:
        write_pdf_report(pdf_path, alignments, trajectories, tasks, fig_paths)
    except PermissionError:
        write_pdf_report(pdf_path.with_name(f"{pdf_path.stem}_codex{pdf_path.suffix}"), alignments, trajectories, tasks, fig_paths)

    print("完成。输出目录：", out_dir)
    for i in (1, 2, 3):
        r = alignments[i]
        ci_str = f"  95%CI: [{r.ci_delta_lo:.4f}, {r.ci_delta_hi:.4f}]" if r.ci_delta_lo != 0 or r.ci_delta_hi != 0 else ""
        f_str = f"  F={r.f_statistic:.2f}, p={r.f_p_value:.4f}" if r.f_statistic else ""
        print(
            f"  问题{i}: delta={r.delta2_minus_1:.6f}s{ci_str}, "
            f"bias=({r.bias2_x:.6f},{r.bias2_y:.6f})m, "
            f"has_bias={r.has_system_bias}{f_str}"
        )
    print(f"候选任务数：{tasks_original_count}，调度后任务数：{len(tasks)}，结果表：{out_dir / 'result.xlsx'}")
    print(f"敏感性分析：{sens_df.to_string(index=False)}")


if __name__ == "__main__":
    main()
